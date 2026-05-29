import openpyxl


def get_sheet_names(xlsx_file):
    '''获取Excel文件的所有工作表名称'''
    return openpyxl.load_workbook(xlsx_file, True).sheetnames


def load_xlsx(xlsx_file, table_name=None):
    '''加载Excel工作表为二维列表'''
    wb = openpyxl.load_workbook(xlsx_file, True)
    if table_name is None:
        ws = wb.active
    else:
        ws = wb[table_name]

    rows = []
    for row in ws:
        rows.append([v.value for v in row])

    return rows


def save_xlsx(rows, xlsx_file):
    '''将二维列表保存为Excel文件'''
    wb = openpyxl.Workbook()
    ws = wb.active

    for row in rows:
        ws.append(row)

    wb.save(xlsx_file)


if __name__ == '__main__':
    print(get_sheet_names(r'tmp_1.xlsx'))
    rows = load_xlsx(r'tmp_1.xlsx')
    save_xlsx(rows, 'tmp_2.xlsx')
